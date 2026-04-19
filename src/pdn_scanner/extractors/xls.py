from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

import xlrd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from pdn_scanner.config import AppConfig
from pdn_scanner.enums import ContentStatus, FileFormat
from pdn_scanner.models import ExtractedContent, FileDescriptor

from .base import BaseExtractor

OLE_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
ZIP_SIGNATURE = b"PK\x03\x04"
HTML_SIGNATURES = (b"<!doctype html", b"<html", b"<?xml")
HEADER_MARKERS = (
    "фамилия",
    "фио",
    "инициалы",
    "full name",
    "first name",
    "last name",
    "middle name",
    "given name",
    "surname",
    "name",
    "email",
    "e-mail",
    "телефон",
    "telephone",
    "phone number",
    "phone",
    "адрес",
    "address",
    "дата рождения",
    "date of birth",
    "birth",
    "birthplace",
    "паспорт",
    "снилс",
    "инн",
    "доход",
    "имущество",
    "объекты недвижимости",
    "вид объекта",
    "площад",
    "страна",
)


class XLSExtractor(BaseExtractor):
    formats = (FileFormat.XLS,)
    name = "xls"

    def extract(self, file_descriptor: FileDescriptor, config: AppConfig) -> ExtractedContent:
        path = Path(file_descriptor.path)
        raw = path.read_bytes()
        warnings: list[str] = []

        try:
            if _looks_like_html(raw):
                chunks, metadata, extra_warnings = _extract_html_table_chunks(raw, config)
                warnings.extend(extra_warnings)
            elif raw.startswith(ZIP_SIGNATURE) or path.suffix.lower() == ".xlsx":
                chunks, metadata = _extract_xlsx_chunks(path, config)
            elif raw.startswith(OLE_SIGNATURE):
                chunks, metadata = _extract_xls_chunks(path, config)
            else:
                warnings.append("Unknown spreadsheet signature; attempting HTML parse fallback")
                chunks, metadata, extra_warnings = _extract_html_table_chunks(raw, config)
                warnings.extend(extra_warnings)
        except Exception as exc:
            return ExtractedContent(
                file_path=file_descriptor.path,
                status=ContentStatus.ERROR,
                text_chunks=[],
                warnings=warnings + [f"Spreadsheet extraction failed: {exc}"],
                metadata={"extractor": self.name},
            )

        status = ContentStatus.OK if chunks else ContentStatus.EMPTY
        metadata["extractor"] = self.name
        return ExtractedContent(
            file_path=file_descriptor.path,
            status=status,
            text_chunks=chunks,
            structured_rows_scanned=int(metadata.get("structured_rows_scanned", len(chunks))),
            sheets_scanned=metadata.get("sheets_scanned"),
            warnings=warnings,
            metadata=metadata,
        )


def _extract_xls_chunks(path: Path, config: AppConfig) -> tuple[list[str], dict]:
    book = xlrd.open_workbook(path)
    sheet_rows: list[tuple[str, list[list[str]]]] = []
    for sheet in book.sheets():
        rows: list[list[str]] = []
        for row_index in range(sheet.nrows):
            values = [_stringify_xls_value(sheet.cell_value(row_index, column_index)) for column_index in range(sheet.ncols)]
            rows.append(values)
        sheet_rows.append((sheet.name, rows))
    return _rows_to_chunks(sheet_rows, config)


def _extract_xlsx_chunks(path: Path, config: AppConfig) -> tuple[list[str], dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_rows: list[tuple[str, list[list[str]]]] = []
        for worksheet in workbook.worksheets:
            rows: list[list[str]] = []
            for row in worksheet.iter_rows(values_only=True):
                rows.append([_stringify_xls_value(value) for value in row])
            sheet_rows.append((worksheet.title, rows))
        return _rows_to_chunks(sheet_rows, config)
    finally:
        workbook.close()


def _extract_html_table_chunks(raw: bytes, config: AppConfig) -> tuple[list[str], dict, list[str]]:
    warnings: list[str] = []
    text = raw.decode("utf-8", errors="ignore")
    lowered = text.lower()
    if "404 not found" in lowered or "an error occurred" in lowered:
        warnings.append("HTML spreadsheet payload looks like an error page; skipped")
        return [], {"header": [], "structured_rows_scanned": 0, "sheets_scanned": 0, "html_payload": True}, warnings

    soup = BeautifulSoup(text, "lxml")
    tables = soup.find_all("table")
    if not tables:
        warnings.append("HTML spreadsheet payload has no tables; skipped")
        return [], {"header": [], "structured_rows_scanned": 0, "sheets_scanned": 0, "html_payload": True}, warnings

    sheet_rows: list[tuple[str, list[list[str]]]] = []
    for index, table in enumerate(tables, start=1):
        rows: list[list[str]] = []
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"])
            values = [" ".join(cell.get_text(" ", strip=True).split()) for cell in cells]
            rows.append(values)
        sheet_rows.append((f"table_{index}", rows))

    chunks, metadata = _rows_to_chunks(sheet_rows, config)
    metadata["html_payload"] = True
    return chunks, metadata, warnings


def _rows_to_chunks(sheet_rows: list[tuple[str, list[list[str]]]], config: AppConfig) -> tuple[list[str], dict]:
    chunks: list[str] = []
    header_names: set[str] = set()
    rows_scanned = 0
    sheets_scanned = 0

    for sheet_name, rows in sheet_rows:
        sheets_scanned += 1
        header_context: list[list[str]] = []
        for row in rows:
            if rows_scanned >= config.detection.max_text_chunks_per_file:
                return chunks, {
                    "header": sorted(header_names),
                    "structured_rows_scanned": rows_scanned,
                    "sheets_scanned": sheets_scanned,
                }

            normalized_row = [_normalize_cell(cell) for cell in row]
            if not any(normalized_row):
                continue

            if _looks_like_header_row(normalized_row):
                header_context.append(normalized_row)
                header_context = header_context[-2:]
                continue

            merged_headers = _merge_headers(header_context, len(normalized_row))
            pieces = [f"sheet: {sheet_name}"]
            non_empty = 0
            for index, value in enumerate(normalized_row):
                if not value:
                    continue
                non_empty += 1
                header = merged_headers[index] if index < len(merged_headers) else f"col_{index + 1}"
                header = _normalize_header_name(header, index)
                header_names.add(header)
                pieces.append(f"{header}: {value}")

            if non_empty == 0:
                continue

            chunks.append(" | ".join(pieces))
            rows_scanned += 1

    return chunks, {
        "header": sorted(header_names),
        "structured_rows_scanned": rows_scanned,
        "sheets_scanned": sheets_scanned,
    }


def _looks_like_html(raw: bytes) -> bool:
    sample = raw[:256].lstrip().lower()
    return any(sample.startswith(signature) for signature in HTML_SIGNATURES)


def _stringify_xls_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_cell(value: str) -> str:
    return " ".join(value.replace("\xa0", " ").replace("\n", " ").split())


def _looks_like_header_row(row: list[str]) -> bool:
    non_empty = [cell for cell in row if cell]
    if len(non_empty) < 2:
        return False
    lowered = " | ".join(non_empty).lower()
    return any(marker in lowered for marker in HEADER_MARKERS)


def _merge_headers(header_rows: list[list[str]], width: int) -> list[str]:
    merged: list[str] = []
    for index in range(width):
        parts: list[str] = []
        for row in header_rows:
            if index >= len(row):
                continue
            value = row[index]
            if value and value not in parts:
                parts.append(value)
        merged.append(" / ".join(parts) if parts else f"col_{index + 1}")
    return merged


def _normalize_header_name(header: str, index: int) -> str:
    lowered = header.lower()
    aliases: tuple[tuple[str, tuple[str, ...]], ...] = (
        ("last name", ("last name", "surname", "family name")),
        ("first name", ("first name", "given name", "forename")),
        ("middle name", ("middle name", "middle initial")),
        ("ФИО", ("фамилия", "фио", "инициалы", "full name", "customer name", "customer_name", "employee name", "employee_name", "name")),
        ("email", ("e-mail", "email", "электронн", "почт")),
        ("телефон", ("телефон", "telephone", "phone number", "phone", "mobile phone", "моб")),
        ("адрес", ("адрес", "mailing address", "home address", "residential address", "residence address", "street address", "address")),
        ("дата рождения", ("дата рождения", "birth date", "date of birth", "date_of_birth", "dob")),
        ("место рождения", ("место рождения", "birth place", "place of birth", "place_of_birth", "birthplace")),
        ("паспорт", ("паспорт", "passport number", "passport no", "passport")),
        ("снилс", ("снилс", "snils")),
        ("инн", ("инн", "inn")),
    )
    for canonical, markers in aliases:
        if any(marker in lowered for marker in markers):
            return canonical
    return header if header else f"col_{index + 1}"
